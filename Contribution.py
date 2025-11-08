import streamlit as st
import pandas as pd
import numpy as np
import io
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import numbers

# =========================
# APP CONFIGURATION
# =========================
st.set_page_config(
    page_title="Granular Spec Multiplier",
    layout="wide",
    initial_sidebar_state="collapsed",
)
st.title("📊 Contribution Matchback of BOSE-US")

st.markdown("""
<style>
@keyframes pulse {
  0% { opacity: 0.3; }
  50% { opacity: 1.0; }
  100% { opacity: 0.3; }
}
.bose-spinner {
  font-size: 28px;
  font-weight: bold;
  color: #003087; /* A deep blue */
  text-align: center;
  padding: 20px;
  animation: pulse 1.5s ease-in-out infinite;
}
</style>
""", unsafe_allow_html=True)

# =========================
# STEP 2 — LOADING HELPERS (MODIFIED FOR STREAMLIT + CACHED)
# =========================

# Cache file loading for performance.
@st.cache_data
def load_main_spec(file_like_object):
    """Auto-detects header row with 'Variable' and 'Type' columns."""
    try:
        xl = pd.ExcelFile(file_like_object)
        sheet_name = next((s for s in xl.sheet_names if "model" in s.lower()), None)
        if not sheet_name:
            st.error("❌ No sheet containing 'Model' found in Main Spec.")
            return None

        preview = pd.read_excel(file_like_object, sheet_name=sheet_name, nrows=50, header=None)
        header_row = None

        for i, row in preview.iterrows():
            cells = [str(x).strip().lower() for x in row.fillna("")]
            short_texts = [c for c in cells if 0 < len(c) <= 20]
            if len(short_texts) < 2:
                continue
            if any("variable" in c for c in short_texts) and any("type" in c for c in short_texts):
                header_row = i
                break
        
        if header_row is None:
            st.error("❌ Could not auto-detect header row (with 'Variable' and 'Type') in Main Spec.")
            return None

        df = pd.read_excel(file_like_object, sheet_name=sheet_name, header=header_row)
        df.columns = [str(c).strip().upper() for c in df.columns]

        possible_var_cols = ["VARIABLE", "VARIABLES", "VARIABLE NAME", "VAR NAME", "VARIABLE_NAME"]
        possible_type_cols = ["TYPE", "TYPES", "VARIABLE TYPE"]

        var_col = next((c for c in df.columns if c in possible_var_cols), None)
        type_col = next((c for c in df.columns if c in possible_type_cols), None)
        if not var_col or not type_col:
            st.error(f"❌ 'Variable' or 'Type' column not found in Main Spec. Found: {df.columns.tolist()}")
            return None

        df = df[[var_col, type_col]].rename(columns={var_col: "VARIABLE", type_col: "TYPE"})
        df["VARIABLE"] = df["VARIABLE"].astype(str).str.strip().str.upper()
        df["TYPE"] = df["TYPE"].astype(str).str.strip().str.title()
        return df
    except Exception as e:
        st.error(f"Error loading Main Spec: {e}")
        return None

@st.cache_data
def load_pmf(file_like_object):
    """Loads PMF sheet and normalizes Geography + Season columns."""
    try:
        xl = pd.ExcelFile(file_like_object)
        sheet_name = next((s for s in xl.sheet_names if "pmf" in s.lower()), None)
        if not sheet_name:
            st.error(f"❌ No sheet named like 'PMF' found. Sheets: {xl.sheet_names}")
            return None
        
        df = pd.read_excel(file_like_object, sheet_name=sheet_name, dtype=str)
        df.columns = [str(c).strip().upper() for c in df.columns]

        geo_col = next((c for c in df.columns if "GEO" in c), None)
        if not geo_col:
            st.error("❌ Geography column not found in PMF.")
            return None
        df.rename(columns={geo_col: "GEOGRAPHY"}, inplace=True)

        possible_season = ["SEASON", "PERIOD MAPPING", "PERIOD_MAPPING", "PERIOD_DEFINITION", "TIME_PERIODS"]
        season_col = next((c for c in df.columns if c in possible_season), None)
        if not season_col:
            st.error("❌ Season column not found in PMF.")
            return None
        df.rename(columns={season_col: "SEASON"}, inplace=True)

        df["GEOGRAPHY"] = df["GEOGRAPHY"].astype(str).str.upper().str.strip()
        df["SEASON"] = df["SEASON"].astype(str).str.upper().str.strip()
        return df
    except Exception as e:
        st.error(f"Error loading PMF: {e}")
        return None

@st.cache_data
def load_granular(file_like_object):
    """Loads MAP sheet and all MAP-related sheets."""
    try:
        xl = pd.ExcelFile(file_like_object)
        map_sheet = next((s for s in xl.sheet_names if "map" in s.lower()), None)
        if not map_sheet:
            st.error("❌ No 'MAP' sheet found in Granular file.")
            return None

        map_df = pd.read_excel(file_like_object, sheet_name=map_sheet, dtype=str)
        map_df.columns = [str(c).strip().upper() for c in map_df.columns]
        
        # Load all sheets
        all_sheets = {s: pd.read_excel(file_like_object, sheet_name=s, dtype=str) for s in xl.sheet_names}
        return map_df, all_sheets
    except Exception as e:
        st.error(f"Error loading Granular Spec: {e}")
        return None

# =========================
# STEP 4 — PREPARE PMF (Refactored to a function)
# =========================
def prepare_pmf_multipliers(pmf_df):
    """Prepares the PMF multiplier dictionary."""
    pmf_vars = [c for c in pmf_df.columns if "_PMF" in c]
    pmf_long = pmf_df.melt(id_vars=["GEOGRAPHY", "SEASON"], value_vars=pmf_vars,
                          var_name="VARIABLE_PMF", value_name="MULTIPLIER")

    pmf_long["VARIABLE"] = pmf_long["VARIABLE_PMF"].str.replace("_PMF", "", regex=False)
    pmf_long["MULTIPLIER"] = pd.to_numeric(pmf_long["MULTIPLIER"], errors="coerce")

    pmf_dict = {(r.GEOGRAPHY, r.SEASON, r.VARIABLE): r.MULTIPLIER for r in pmf_long.itertuples()}
    return pmf_dict

# =========================
# STEP 5 — APPLY MULTIPLIERS (Refactored to a function)
# =========================
def normalize_geo(name: str):
    """Treat BOSE.COM, BOSE_COM, BOSE COM as identical."""
    return str(name).strip().upper().replace(".", "").replace("_", "").replace(" ", "")

def find_multiplier_for(sheet_name, season, var, mapcode_to_geo, pmf_dict_u):
    """Find multiplier using map mapping and fallback logic."""
    sheet_up = normalize_geo(sheet_name)
    season_up = str(season).strip().upper()
    var_up = str(var).strip().upper()

    geo = mapcode_to_geo.get(sheet_up)
    if geo:
        m = pmf_dict_u.get((normalize_geo(geo), season_up, var_up))
        if m is not None and not pd.isna(m):
            return m, geo

    m = pmf_dict_u.get((sheet_up, season_up, var_up))
    if m is not None and not pd.isna(m):
        return m, sheet_up

    return None, None

def apply_multipliers(granular_sheets, map_df, pmf_dict, selected_vars):
    """Main processing function to apply multipliers."""
    # Build reverse map: MAP_CODE -> GEOGRAPHY
    map_df_copy = map_df.copy()
    map_df_copy["GEOGRAPHY"] = map_df_copy["GEOGRAPHY"].astype(str).str.strip().str.upper()
    map_df_copy["MAP"] = map_df_copy["MAP"].astype(str).str.strip().str.upper()
    mapcode_to_geo = dict(zip(map_df_copy["MAP"].tolist(), map_df_copy["GEOGRAPHY"].tolist()))

    # PMF dict normalized for comparison
    pmf_dict_u = {
        (normalize_geo(k[0]), str(k[1]).strip().upper(), str(k[2]).strip().upper()): v
        for k, v in pmf_dict.items()
    }

    selected_vars_set = set([v.strip().upper() for v in selected_vars])
    multiplied_records, skipped_records = [], []
    updated_sheets = {}

    for sheet_name, df_full in granular_sheets.items():
        df = df_full.copy()
        cols_upper = [str(c).upper() for c in df.columns]

        if not {"VARIABLE", "CONTRIBUTION", "MIN", "MAX"}.issubset(cols_upper):
            updated_sheets[sheet_name] = df
            continue

        var_col = df.columns[cols_upper.index("VARIABLE")]
        contrib_col = df.columns[cols_upper.index("CONTRIBUTION")]
        min_col = df.columns[cols_upper.index("MIN")]
        max_col = df.columns[cols_upper.index("MAX")]

        # Convert MIN/MAX columns to numeric safely before multiplication
        for colname in [min_col, max_col]:
            df[colname] = (
                df[colname]
                .astype(str)
                .str.replace("%", "", regex=False)
                .str.replace(",", "", regex=False)
                .apply(lambda x: pd.to_numeric(x, errors="coerce") if pd.notna(x) else np.nan)
            )

        for idx in df.index:
            raw_var = df.at[idx, var_col]
            raw_season = df.at[idx, contrib_col]

            if pd.isna(raw_var) or str(raw_var).strip() == "":
                continue

            var = str(raw_var).strip().upper()
            season = str(raw_season).strip().upper() if not pd.isna(raw_season) else ""

            if var not in selected_vars_set:
                skipped_records.append([sheet_name, var, season, "VAR_NOT_SELECTED", None, None, None, None])
                continue
            if season == "" or season.upper() in ["NAN", "NONE"]:
                skipped_records.append([sheet_name, var, season, "NO_SEASON", None, None, None, None])
                continue

            multiplier, used_geo = find_multiplier_for(sheet_name, season, var, mapcode_to_geo, pmf_dict_u)
            if multiplier is None or pd.isna(multiplier):
                skipped_records.append([sheet_name, var, season, "NO_MULTIPLIER_FOUND", None, None, None, None])
                continue

            try:
                old_min = float(df.at[idx, min_col]) if not pd.isna(df.at[idx, min_col]) else None
                old_max = float(df.at[idx, max_col]) if not pd.isna(df.at[idx, max_col]) else None
            except:
                skipped_records.append([sheet_name, var, season, "MIN_MAX_NOT_NUMERIC", None, None, None, None])
                continue

            if old_min is None or old_max is None:
                continue

            new_min = round(old_min * float(multiplier), 6)
            new_max = round(old_max * float(multiplier), 6)

            df.at[idx, min_col] = new_min
            df.at[idx, max_col] = new_max

            multiplied_records.append([sheet_name, var, season, used_geo, multiplier, old_min, old_max, new_min, new_max])

        updated_sheets[sheet_name] = df

    multiplied_df = pd.DataFrame(
        multiplied_records,
        columns=["MAP_SHEET", "VARIABLE", "SEASON", "PMF_GEOGRAPHY_USED",
                 "MULTIPLIER", "OLD_MIN", "OLD_MAX", "NEW_MIN", "NEW_MAX"]
    )

    skipped_df = pd.DataFrame(
        skipped_records,
        columns=["MAP_SHEET", "VARIABLE", "SEASON", "REASON",
                 "OLD_MIN", "OLD_MAX", "NEW_MIN", "NEW_MAX"]
    )
    
    return updated_sheets, multiplied_df, skipped_df

# =========================
# STEP 6 — SAVE OUTPUT (Refactored to return in-memory files)
# =========================
def create_output_excel(granular_sheets, updated_sheets):
    """Saves all sheets to an in-memory Excel file and applies formatting."""
    output_buffer = io.BytesIO()

    with pd.ExcelWriter(output_buffer, engine="openpyxl") as writer:
        for sheet_name, df in granular_sheets.items():
            # Use the updated df if it exists, otherwise use the original
            df_to_save = updated_sheets.get(sheet_name, df)
            df_to_save.to_excel(writer, sheet_name=sheet_name, index=False)
    
    # Apply % formatting
    # We must seek(0) to "rewind" the buffer for openpyxl to read it
    output_buffer.seek(0)
    wb = load_workbook(output_buffer)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row <= 1: # Skip empty or header-only sheets
             continue
             
        header = [cell.value for cell in ws[1] if cell.value]

        min_idx = max_idx = None
        for i, h in enumerate(header, start=1):
            if str(h).strip().upper() == "MIN":
                min_idx = i
            elif str(h).strip().upper() == "MAX":
                max_idx = i
        
        if min_idx is None and max_idx is None:
            continue # No MIN/MAX columns found in this sheet

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if min_idx:
                c = row[min_idx - 1]
                if isinstance(c.value, (int, float)):
                    c.number_format = numbers.FORMAT_PERCENTAGE_00
            if max_idx:
                c = row[max_idx - 1]
                if isinstance(c.value, (int, float)):
                    c.number_format = numbers.FORMAT_PERCENTAGE_00

    # Save formatted workbook back to a new buffer
    final_buffer = io.BytesIO()
    wb.save(final_buffer)
    wb.close()
    
    return final_buffer.getvalue()

def create_log_excel(multiplied_df, skipped_df):
    """Creates the in-memory log file."""
    log_buffer = io.BytesIO()
    
    mult_summary = (
        multiplied_df.groupby("MAP_SHEET").size().reset_index(name="MULTIPLIED_COUNT")
        if not multiplied_df.empty else pd.DataFrame(columns=["MAP_SHEET","MULTIPLIED_COUNT"])
    )
    skip_summary = (
        skipped_df.groupby(["MAP_SHEET","REASON"]).size().reset_index(name="SKIPPED_COUNT")
        if not skipped_df.empty else pd.DataFrame(columns=["MAP_SHEET","REASON","SKIPPED_COUNT"])
    )
    
    summary_df = mult_summary.merge(skip_summary, on="MAP_SHEET", how="outer").fillna(0)
    
    if not summary_df.empty:
        summary_df.loc["Total"] = {
            "MAP_SHEET": "GRAND TOTAL",
            "MULTIPLIED_COUNT": mult_summary["MULTIPLIED_COUNT"].sum() if not mult_summary.empty else 0,
            "SKIPPED_COUNT": skip_summary["SKIPPED_COUNT"].sum() if not skip_summary.empty else 0
        }
    
    with pd.ExcelWriter(log_buffer, engine="openpyxl") as writer:
        multiplied_df.to_excel(writer, sheet_name="Multiplied", index=False)
        skipped_df.to_excel(writer, sheet_name="Skipped", index=False)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
    
    return log_buffer.getvalue(), summary_df

# =========================
# STREAMLIT UI - ORGANIZED BY TABS
# =========================

# Initialize session state keys
defaults = {
    "step1_complete": False,
    "step2_complete": False,
    "step3_complete": False,
    "main_spec": None,
    "pmf": None,
    "map_df": None,
    "granular_sheets": None,
    "selected_types": None,
    "selected_vars": None,
    "pmf_dict": None,
    "updated_sheets": None,
    "multiplied_df": None,
    "skipped_df": None,
    "output_file_bytes": None,
    "log_file_bytes": None,
    "log_summary_df": None,
}
for key, value in defaults.items():
    if key not in st.session_state:
        st.session_state[key] = value


tab1, tab2, tab3, tab4 = st.tabs([
    "Step 1: Upload Files", 
    "Step 2: Select Types", 
    "Step 3: Run Process", 
    "Step 4: Download Results"
])

# =========================
# TAB 1: FILE UPLOAD (Original Step 1 & 2)
# =========================
with tab1:
    st.header("Step 1: Upload Files")
    st.info("Please upload the three required Excel files. Processing will begin automatically.")

    gran_file = st.file_uploader("1. Select Granular Spec (Excel)", type=['xlsx'])
    pmf_file = st.file_uploader("2. Select PMF File (Excel)", type=['xlsx'])
    main_file = st.file_uploader("3. Select Main Spec (Excel)", type=['xlsx'])

    if gran_file and pmf_file and main_file:
        spinner_placeholder = st.empty()
        spinner_placeholder.markdown('<div class="bose-spinner">BOSE</div>', unsafe_allow_html=True)
        
        # Load all files
        main_spec_df = load_main_spec(main_file)
        pmf_df = load_pmf(pmf_file)
        map_df, granular_sheets_dict = load_granular(gran_file)

        # Check if all loaded successfully
        if main_spec_df is not None and pmf_df is not None and map_df is not None:
            spinner_placeholder.empty() # Clear spinner
            st.session_state["main_spec"] = main_spec_df
            st.session_state["pmf"] = pmf_df
            st.session_state["map_df"] = map_df
            st.session_state["granular_sheets"] = granular_sheets_dict
            st.session_state["step1_complete"] = True
                
            st.success("✅ All files loaded and validated successfully!")
            st.subheader("File Previews (First 5 Rows)")
            st.write("**Main Spec (Processed)**")
            st.dataframe(main_spec_df.head())
            st.write("**PMF (Processed)**")
            st.dataframe(pmf_df.head())
            st.write("**Granular MAP (Processed)**")
            st.dataframe(map_df.head())
        else:
            spinner_placeholder.empty() # Clear spinner
            st.error("One or more files failed to load. Please check errors above.")
            st.session_state["step1_complete"] = False

# =========================
# TAB 2: TYPE SELECTION (Original Step 3)
# =========================
with tab2:
    st.header("Step 2: Select Variable Types")
    if not st.session_state["step1_complete"]:
        st.warning("Please upload all files in Step 1 first.")
    else:
        types_available = sorted(st.session_state["main_spec"]["TYPE"].dropna().unique().tolist())
        
        choice_map = {
            "Base": ["Base"],
            "Incremental": ["Incremental"],
            "Base and Incremental": ["Base", "Incremental"]
        }
        
        type_choice = st.radio(
            "Choose variable types to process:",
            options=choice_map.keys(),
            index=2 # Default to "Base and Incremental"
        )
        
        selected_types = choice_map[type_choice]
        selected_vars = st.session_state["main_spec"][
            st.session_state["main_spec"]["TYPE"].isin(selected_types)
        ]["VARIABLE"].tolist()
        
        st.session_state["selected_types"] = selected_types
        st.session_state["selected_vars"] = selected_vars
        st.session_state["step2_complete"] = True

        st.success(f"✅ Selected types: **{', '.join(selected_types)}**")
        st.info(f"Total variables to process: **{len(selected_vars)}**")
        
        with st.expander("Click to see all selected variables"):
            st.dataframe(selected_vars)

# =========================
# TAB 3: RUN PROCESS (Original Step 4 & 5)
# =========================
with tab3:
    st.header("Step 3: Apply Multipliers")
    if not st.session_state["step2_complete"]:
        st.warning("Please complete Steps 1 and 2 first.")
    else:
        st.info("This step will prepare the PMF multipliers and apply them to all sheets.")
        
        if st.button("🚀 Run Multiplier Process", type="primary", use_container_width=True):
            spinner_placeholder = st.empty()
            spinner_placeholder.markdown('<div class="bose-spinner">BOSE</div>', unsafe_allow_html=True)
            
            try:
                # Step 4: Prepare PMF
                pmf_dict = prepare_pmf_multipliers(st.session_state["pmf"])
                st.session_state["pmf_dict"] = pmf_dict
                
                # Step 5: Apply Multipliers
                updated_sheets, multiplied_df, skipped_df = apply_multipliers(
                    st.session_state["granular_sheets"],
                    st.session_state["map_df"],
                    st.session_state["pmf_dict"],
                    st.session_state["selected_vars"]
                )
                
                st.session_state["updated_sheets"] = updated_sheets
                st.session_state["multiplied_df"] = multiplied_df
                st.session_state["skipped_df"] = skipped_df
                st.session_state["step3_complete"] = True
                
                spinner_placeholder.empty() # Clear spinner
                st.success("✅ Multipliers applied successfully!")
                st.metric("Records Multiplied", len(multiplied_df))
                st.metric("Records Skipped", len(skipped_df))
                
            except Exception as e:
                spinner_placeholder.empty() # Clear spinner
                st.error(f"An error occurred during processing: {e}")
                st.session_state["step3_complete"] = False

# =========================
# TAB 4: DOWNLOAD RESULTS (Original Step 6)
# =========================
with tab4:
    st.header("Step 4: Download Results")
    if not st.session_state["step3_complete"]:
        st.warning("Please run the process in Step 3 first.")
    else:
        st.success("🎉 Process completed! Your files are ready for download.")
        
        # --- Generate files ---
        if st.session_state["output_file_bytes"] is None:
            spinner_placeholder_1 = st.empty()
            spinner_placeholder_1.markdown('<div class="bose-spinner">BOSE</div>', unsafe_allow_html=True)
            output_bytes = create_output_excel(
                st.session_state["granular_sheets"],
                st.session_state["updated_sheets"]
            )
            st.session_state["output_file_bytes"] = output_bytes
            spinner_placeholder_1.empty()
        
        # --- Create file names ---
        today_str = date.today().isoformat()
        types_str = '_'.join(st.session_state['selected_types'])
        granular_outfile_name = f"Granular_Updated_{types_str}_{today_str}.xlsx"
        log_outfile_name = f"Granular_Log_{types_str}_{today_str}.xlsx"

        # --- Download Button for Main File ---
        st.download_button(
            label="💾 Download Updated Granular File",
            data=st.session_state["output_file_bytes"],
            file_name=granular_outfile_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        
        st.markdown("---")
        
        # --- Log File Generation ---
        generate_log = st.checkbox("Generate detailed log file?", value=True)
        
        if generate_log:
            if st.session_state["log_file_bytes"] is None:
                spinner_placeholder_2 = st.empty()
                spinner_placeholder_2.markdown('<div class="bose-spinner">BOSE</div>', unsafe_allow_html=True)
                log_bytes, summary_df = create_log_excel(
                    st.session_state["multiplied_df"],
                    st.session_state["skipped_df"]
                )
                st.session_state["log_file_bytes"] = log_bytes
                st.session_state["log_summary_df"] = summary_df
                spinner_placeholder_2.empty()

            if st.session_state["log_file_bytes"]:
                # --- Download Button for Log File ---
                st.download_button(
                    label="📋 Download Log File",
                    data=st.session_state["log_file_bytes"],
                    file_name=log_outfile_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
                st.subheader("Log Summary")
                st.dataframe(st.session_state["log_summary_df"])

        # --- Show full log details in expanders ---
        with st.expander("View Multiplied Records Details"):
            st.dataframe(st.session_state["multiplied_df"])
        
        with st.expander("View Skipped Records Details"):
            st.dataframe(st.session_state["skipped_df"])